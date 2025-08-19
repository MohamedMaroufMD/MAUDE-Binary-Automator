#!/usr/bin/env python3
"""
MAUDE Binary Columns Automator
==============================

This script automatically adds binary columns to MAUDE Excel files for:
- Device problems (all unique values)
- Patient problems (all unique values) 
- Patient outcomes (all unique values)

Features:
- Preserves all original styling, fonts, colors, and formatting
- Automatically detects unique problems and outcomes
- Applies green/red color coding for binary values
- Works with any MAUDE file that has an 'Events' sheet
- Handles different file names and structures

Usage:
    python3 maude_binary_automator.py [excel_file_path]
    
    If no file path is provided, the script will look for Excel files in the current directory.
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import os
import sys
import glob
from datetime import datetime

def find_maude_files():
    """Find MAUDE Excel files in the current directory."""
    excel_files = glob.glob("*.xlsx")
    maude_files = [f for f in excel_files if "MAUDE" in f.upper()]
    return maude_files

def validate_maude_file(file_path):
    """Validate that the file has the required structure."""
    try:
        # Check if file exists
        if not os.path.exists(file_path):
            return False, f"File not found: {file_path}"
        
        # Check if it's an Excel file
        if not file_path.lower().endswith('.xlsx'):
            return False, f"Not an Excel file: {file_path}"
        
        # Check if it has an Events sheet
        xl = pd.ExcelFile(file_path)
        if 'Events' not in xl.sheet_names:
            return False, f"No 'Events' sheet found in {file_path}"
        
        # Check if it has the required columns
        df = pd.read_excel(file_path, sheet_name='Events')
        device_cols = [col for col in df.columns if 'Device Problem' in col]
        patient_problem_cols = [col for col in df.columns if 'Patient Problem' in col]
        patient_outcome_cols = [col for col in df.columns if 'Patient Outcome' in col]
        
        if not device_cols and not patient_problem_cols and not patient_outcome_cols:
            return False, f"No problem/outcome columns found in {file_path}"
        
        return True, "File validated successfully"
        
    except Exception as e:
        return False, f"Error validating file: {str(e)}"

def check_existing_binary_columns(df):
    """Check for existing binary columns and return them."""
    existing_binary = []
    for col in df.columns:
        if (col.startswith('Device_') or col.startswith('Patient_') or col.startswith('Outcome_')) and \
           col not in [c for c in df.columns if 'Device Problem' in c or 'Patient Problem' in c or 'Patient Outcome' in c]:
            existing_binary.append(col)
    return existing_binary

def create_binary_columns(df, device_values, patient_problem_values, patient_outcome_values, 
                         device_cols, patient_problem_cols, patient_outcome_cols, existing_binary_cols):
    """Create binary columns efficiently using vectorized operations, avoiding duplicates."""
    binary_data = {}
    
    # Create binary columns for device problems
    for problem in device_values:
        col_name = f"Device_{problem.replace(' ', '_').replace(',', '').replace('(', '').replace(')', '').replace('/', '_')}"
        if col_name not in existing_binary_cols:
            # Use vectorized operations to check across all device problem columns
            binary_data[col_name] = df[device_cols].eq(problem).any(axis=1).astype(int)
    
    # Create binary columns for patient problems
    for problem in patient_problem_values:
        col_name = f"Patient_{problem.replace(' ', '_').replace(',', '').replace('(', '').replace(')', '').replace('/', '_')}"
        if col_name not in existing_binary_cols:
            binary_data[col_name] = df[patient_problem_cols].eq(problem).any(axis=1).astype(int)
    
    # Create binary columns for patient outcomes
    for outcome in patient_outcome_values:
        col_name = f"Outcome_{outcome.replace(' ', '_').replace(',', '').replace('(', '').replace(')', '').replace('/', '_')}"
        if col_name not in existing_binary_cols:
            binary_data[col_name] = df[patient_outcome_cols].eq(outcome).any(axis=1).astype(int)
    
    return binary_data

def capture_original_styling(ws):
    """Capture all original styling from the worksheet."""
    original_styles = {}
    original_row_heights = {}
    original_col_widths = {}
    
    # Capture original formatting for each cell
    for row in range(1, ws.max_row + 1):
        original_row_heights[row] = ws.row_dimensions[row].height
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            original_styles[(row, col)] = {
                'font': copy(cell.font),
                'fill': copy(cell.fill),
                'alignment': copy(cell.alignment),
                'border': copy(cell.border),
                'number_format': cell.number_format
            }
    
    # Capture column widths
    for col_letter in ws.column_dimensions:
        original_col_widths[col_letter] = ws.column_dimensions[col_letter].width
    
    return original_styles, original_row_heights, original_col_widths

def get_unique_values(df, column_patterns):
    """Get unique values from columns matching the patterns."""
    unique_values = set()
    matching_cols = []
    
    for pattern in column_patterns:
        cols = [col for col in df.columns if pattern in col]
        matching_cols.extend(cols)
        for col in cols:
            unique_values.update(df[col].dropna().unique())
    
    return sorted(list(unique_values)), matching_cols

def apply_binary_formatting(ws, df_with_binary, original_col_count, binary_data):
    """Apply green/red formatting to binary columns."""
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    
    binary_start_col = original_col_count + 1
    binary_cols = list(binary_data.keys())
    
    # Format binary columns
    for col_idx, col_name in enumerate(binary_cols, start=binary_start_col):
        # Format header
        header_cell = ws.cell(row=1, column=col_idx)
        header_cell.font = Font(bold=True)
        
        # Format data cells
        for row_idx in range(2, len(df_with_binary) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value == 1:
                cell.fill = green_fill
            elif cell.value == 0:
                cell.fill = red_fill

def restore_original_formatting(ws, original_styles, original_row_heights, original_col_widths, original_col_count):
    """Restore original formatting for existing columns."""
    for row in range(1, ws.max_row + 1):
        # Set row height if it was originally set
        if row in original_row_heights and original_row_heights[row]:
            ws.row_dimensions[row].height = original_row_heights[row]
            
        for col in range(1, original_col_count + 1):
            if (row, col) in original_styles:
                cell = ws.cell(row=row, column=col)
                style = original_styles[(row, col)]
                cell.font = style['font']
                cell.fill = style['fill']
                cell.alignment = style['alignment']
                cell.border = style['border']
                cell.number_format = style['number_format']
    
    # Apply column widths for original columns
    for col_letter, width in original_col_widths.items():
        if width:
            ws.column_dimensions[col_letter].width = width

def process_maude_file(file_path):
    """Main function to process a MAUDE Excel file."""
    print(f"\n{'='*60}")
    print(f"Processing: {file_path}")
    print(f"{'='*60}")
    
    # Validate the file
    is_valid, message = validate_maude_file(file_path)
    if not is_valid:
        print(f"❌ {message}")
        return False
    
    print(f"✅ {message}")
    
    try:
        # Read the Events sheet data
        df = pd.read_excel(file_path, sheet_name='Events')
        print(f"📊 Loaded {len(df)} rows and {len(df.columns)} columns")
        
        # Load the workbook to access styling
        book = load_workbook(file_path)
        ws = book['Events']
        
        print("🎨 Capturing original styling...")
        original_styles, original_row_heights, original_col_widths = capture_original_styling(ws)
        print("✅ Original styling captured successfully!")
        
        # Get unique values for each category
        device_values, device_cols = get_unique_values(df, ['Device Problem'])
        patient_problem_values, patient_problem_cols = get_unique_values(df, ['Patient Problem'])
        patient_outcome_values, patient_outcome_cols = get_unique_values(df, ['Patient Outcome'])
        
        print(f"\n📋 Found unique values:")
        print(f"   • Device Problems: {len(device_values)}")
        print(f"   • Patient Problems: {len(patient_problem_values)}")
        print(f"   • Patient Outcomes: {len(patient_outcome_values)}")
        
        if not device_values and not patient_problem_values and not patient_outcome_values:
            print("❌ No problems or outcomes found to create binary columns for")
            return False
        
        # Check for existing binary columns
        existing_binary_cols = check_existing_binary_columns(df)
        print(f"\n📋 Found existing binary columns: {len(existing_binary_cols)}")
        for col in existing_binary_cols:
            print(f"   • {col}")
        
        # Create binary columns
        print("\n🔧 Creating binary columns...")
        binary_data = create_binary_columns(df, device_values, patient_problem_values, patient_outcome_values,
                                          device_cols, patient_problem_cols, patient_outcome_cols, existing_binary_cols)
        
        if not binary_data:
            print("✅ No new binary columns needed - all problems/outcomes already have binary columns!")
            return True
        
        # Add binary columns to dataframe
        binary_df = pd.DataFrame(binary_data)
        df_with_binary = pd.concat([df, binary_df], axis=1)
        
        print(f"✅ {len(binary_data)} new binary columns created. Total columns now: {len(df_with_binary.columns)}")
        
        # Clear the worksheet but preserve the workbook structure
        ws.delete_rows(1, ws.max_row)
        ws.delete_cols(1, ws.max_column)
        
        # Write the new data
        for r in dataframe_to_rows(df_with_binary, index=False, header=True):
            ws.append(r)
        
        print("📝 Data written. Now restoring original formatting...")
        
        # Restore original formatting
        original_col_count = len(df.columns)
        restore_original_formatting(ws, original_styles, original_row_heights, original_col_widths, original_col_count)
        print("✅ Original formatting restored!")
        
        # Apply binary column formatting
        print("🎨 Applying green/red formatting to binary columns...")
        apply_binary_formatting(ws, df_with_binary, original_col_count, binary_data)
        print("✅ Binary column formatting applied!")
        
        # Create backup filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = file_path.replace('.xlsx', f'_backup_{timestamp}.xlsx')
        
        # Save backup first
        book.save(backup_path)
        print(f"💾 Backup saved: {backup_path}")
        
        # Save the modified file
        book.save(file_path)
        
        print(f"\n🎉 Successfully processed {file_path}!")
        print(f"   • Original columns: {original_col_count}")
        print(f"   • Binary columns added: {len(binary_data)}")
        print(f"   • Total columns: {len(df_with_binary.columns)}")
        print(f"   • Total rows: {len(df_with_binary)}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error processing file: {str(e)}")
        return False

def main():
    """Main function to handle command line arguments and file processing."""
    print("🚀 MAUDE Binary Columns Automator")
    print("=" * 40)
    
    # Get file path from command line argument or find files
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        if not os.path.exists(file_path):
            print(f"❌ File not found: {file_path}")
            return
        files_to_process = [file_path]
    else:
        # Find MAUDE files in current directory
        files_to_process = find_maude_files()
        if not files_to_process:
            print("❌ No MAUDE Excel files found in current directory")
            print("   Please provide a file path as an argument or place MAUDE files in this directory")
            return
    
    print(f"📁 Found {len(files_to_process)} file(s) to process")
    
    # Process each file
    successful = 0
    for file_path in files_to_process:
        if process_maude_file(file_path):
            successful += 1
    
    print(f"\n{'='*60}")
    print(f"📊 Processing Summary:")
    print(f"   • Files processed: {len(files_to_process)}")
    print(f"   • Successful: {successful}")
    print(f"   • Failed: {len(files_to_process) - successful}")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()
